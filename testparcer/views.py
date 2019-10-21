from datetime import datetime
from datetime import timedelta

from celery.result import AsyncResult
from openpyxl import Workbook
from django.http import HttpResponse, JsonResponse
from django.http import HttpResponse
from django.shortcuts import render, redirect
from testparcer.models import Parceline
from django.db.models import Count, Q, Sum, F
from django.db import connection
from django.core import serializers
from celery import task




def index(request):
    return render(request, 'index.html')


def stats(request, ip):
    return render(request, 'stats.html', {'ip': ip})


def get_data_by_year(request, ip):
    queryset = Parceline.objects.filter(ipaddr=ip)
    y2019 = queryset.filter(dtimefield__year=2019).count()
    y2018 = queryset.filter(dtimefield__year=2018).count()
    y2017 = queryset.filter(dtimefield__year=2017).count()
    y2016 = queryset.filter(dtimefield__year=2016).count()
    y2015 = queryset.filter(dtimefield__year=2015).count()
    json = {'2019': y2019, '2018': y2018, '2017': y2017, '2016': y2016, '2015': y2015}
    return JsonResponse(json)


def common_table(request):
    if request.method == 'GET':
        allcount = Parceline.objects.all().count()
        queryset = Parceline.objects.all()[:20]
        return render(request, 'common_table.html', {'table': queryset, 'allcount': allcount})
    return render(request, 'common_table.html')


def common_table_p(request, page):
    if request.method == 'GET':
        allcount = Parceline.objects.all().count()
        if page < 1:
            page = 1
            return render(request, 'common_table.html', {'allcount': allcount})
        if page < ((allcount - (allcount % 20))/20) - 1:
            queryset = Parceline.objects.all()[20*(page-1):20*page]
            pages = {'prev': page - 1, 'curr': page, 'next': page + 1}
            return render(request, 'common_table.html', {'table': queryset, 'allcount': allcount, })
        else:
            pages = {'prev': page-1, 'curr': page, 'next': page+1}
            queryset = Parceline.objects.all()[20 * (page - 1):]
            return render(request, 'common_table.html', {'table': queryset, 'allcount': allcount, })


def top10(request):
    if request.method == 'GET':
        stats = Parceline.objects.raw('SELECT 1 as id, ' +
                                       "testparcer_parceline.ipaddr, Count(testparcer_parceline.ipaddr) AS total, Count(DISTINCT testparcer_parceline.httpmethod) AS unique_methods, " +
                                       "Sum(REGEXP_REPLACE(COALESCE(testparcer_parceline.bytesread, '0'), '[^0-9]*' ,'0')::integer) AS summ FROM " +
                                       "testparcer_parceline GROUP BY testparcer_parceline.ipaddr ORDER BY summ DESC LIMIT 10")
        return render(request, 'top10.html', {'stats': stats})
    return render(request, 'top10.html')


def xlssave(request):
    parceline_queryset = Parceline.objects.all()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename={date}-parsedlines.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'),)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'access.log'
    columns = [
    'ipaddr',
    'dtimefield',
    'httpmethod',
    'urlrequest',
    'responsecode',
    'bytesread',
    'referrer',
    'user_agent'
    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    for Line in parceline_queryset:
        row_num += 1
        row = [
            Line.ipaddr,
            Line.dtimefield,
            Line.httpmethod,
            Line.urlrequest,
            Line.responsecode,
            Line.bytesread,
            Line.referrer,
            Line.user_agent
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response


def aggregated(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename={date}-parsedlines.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'),)
    stats = Parceline.objects.raw('SELECT 1 as id, ' +
                                  "testparcer_parceline.ipaddr, Count(testparcer_parceline.ipaddr) AS total, Count(DISTINCT testparcer_parceline.httpmethod) AS unique_methods, " +
                                  "Sum(REGEXP_REPLACE(COALESCE(testparcer_parceline.bytesread, '0'), '[^0-9]*' ,'0')::integer) AS summ FROM " +
                                  "testparcer_parceline GROUP BY testparcer_parceline.ipaddr ORDER BY summ DESC LIMIT 10")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'access.log'
    columns = [
    'ip',
    'total',
    'unique Methods',
    'sum of read bytes',
    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    for Line in stats:
        row_num += 1
        row = [
            Line.ipaddr,
            Line.total,
            Line.unique_methods,
            str(Line.summ)
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response


def get_progress(request, task_id):
    result = AsyncResult(task_id)
    response_data = {
        'state': result.state,
        'details': result.info,
    }
    return JsonResponse(response_data)